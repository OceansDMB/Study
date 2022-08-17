from dataclasses import asdict
from lib2to3.pgen2 import driver
from msilib.schema import File
from tkinter import BROWSE
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import re
from openpyxl import load_workbook
import time
import datetime
import pyautogui


# 저장 영역 및 참조 데이터 불러오기(추후 수정)
fpath = r"C:\Users\user\Documents\Study\2.Python\WebCrawler\Excel\data\Crawling_data.xlsx"
dbpath = r"C:\Users\user\Documents\Study\2.Python\WebCrawler\Excel\data\sidogun.xlsx"

# 1) 저장영역 참조 데이터 변수선언
wb = openpyxl.load_workbook(fpath)
now = datetime.date.today
databook = openpyxl.load_workbook(dbpath)
databookWs = databook.active
crawling_data = wb.active

# 크롬 브라우저 패치, 설정변수 전역 초기화, 크롤링 세팅값 지정
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

service = Service(executable_path=ChromeDriverManager().install())
browser = webdriver.Chrome(service=service, options=chrome_options)


# 검색받을 데이터값 할당받기
keyword = pyautogui.prompt("검색어를 입력해 주세요")

# 화면 스크롤 초기값 할당
beforeH = browser.execute_script("return window.scrollY")

# 브라우저 오픈, 크롤링 시작점.
browser.get("https://map.naver.com/")
browser.implicitly_wait(1)

# 브라우저 서치 css값 search 변수에 할당
search = browser.find_element(
    By.CSS_SELECTOR, 'input.input_search')
search.click()

# 브라우저 검색창 클리어 css값 clear 변수에 할당
# x 버튼 활성화
search.send_keys(" ")
clearSearch = browser.find_element(
    By.CSS_SELECTOR, 'button.button_clear')

# 데이터 크롤링 관련 지역 할당.
all_values = []
for row in databookWs.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
        all_values.append(row_value)
# print(all_values)

i = 0
for crawling in all_values:
    browser.implicitly_wait(1)
    clearSearch = browser.find_element(
        By.CSS_SELECTOR, 'button.button_clear')
    clearSearch.click()
    time.sleep(1)
    search.send_keys(all_values[i])
    time.sleep(1)
    search.send_keys(Keys.ENTER)
    time.sleep(3)
    clearSearch = browser.find_element(
        By.CSS_SELECTOR, 'button.button_clear')
    clearSearch.click()
    search.send_keys(keyword)
    time.sleep(1)
    search.send_keys(Keys.ENTER)
    time.sleep(1)
    frame = browser.find_element(By.CSS_SELECTOR, "iframe#searchIframe")
    browser.switch_to.frame(frame)
    time.sleep(1)
    scroll_div = browser.find_element(
        By.XPATH, "/html/body/div[3]/div/div/div[1]")
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(1)
    browser.execute_script("arguments[0].scrollBy(0,2000);", scroll_div)
    time.sleep(1)
    browser.execute_script("arguments[0].scrollBy(0,2000);", scroll_div)
    time.sleep(1)
    browser.execute_script("arguments[0].scrollBy(0,2000);", scroll_div)
    time.sleep(1)
    browser.execute_script("arguments[0].scrollBy(0,2000);", scroll_div)
    time.sleep(1)
    # 여기까지 scroll
    # 맨 아래까지 내려서 해당 페이지의 내용이 다 표시되게 함
    final_result = []
    j = 2
    while j <= 5:
        stores_box = browser.find_element(
            By.XPATH, "/html/body/div[3]/div/div/div[1]/ul")
        stores = browser.find_elements(
            By.XPATH, "/html/body/div[3]/div/div/div[1]/ul/li")
        print(stores)
        # 해당 페이지에서 표시된 모든 가게 정보
        k = 1
        for store in stores:
            name = store.find_element(
                By.CSS_SELECTOR, f"#_pcmap_list_scroll_container > ul > li:nth-child({k}) > div._3ZU00> a:nth-child(1) > div > div > span.place_bluelink._3Apve").text  # 업체명 크롤링 시작점
            click_name = store.find_element(By.CSS_SELECTOR, "div._2w9xx")
            click_name.click()
           #  browser.switch_to.default_content()  # 브라우저 내부 세부 프레임 전환토록 default frame 전환
            time.sleep(7)
            frame_in = browser.find_element(
                By.CSS_SELECTOR, "#app-root > div > div > div > div:nth-child(6) > div > div.place_section.no_margin._18vYz")
            browser.switch_to.frame(frame_in)  # 세부프레임으로 focus 진입
            time.sleep(3)
            try:
                link_url = browser.find_element_by_css_selector(
                    "a._1RUzg").text
            except:
                link_url = " "
            store_info = {
                '업체명': name,
                '주소': link_url
            }
            # Crawling data 를 store_info 변수에 저장
            print(store_info)
            final_result.append(store_info)
            browser.switch_to.default_content()
            browser.switch_to.frame(frame)
            time.sleep(8)
            # 한 페이지 크롤링 끝
            next_button = browser.find_element_by_link_text(str(j))
            next_button.click()
            j = j+1
            k = k+1
            time.sleep(8)
i = i+1

print(final_result)
