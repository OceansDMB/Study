import openpyxl

# 0) define file path

fpath = r"C:\Users\user\Documents\Study\2.Python\WebCrawler\Excel\data\Crawling_data.xlsx"

# 1) import xl
wb = openpyxl.load_workbook(fpath)

# 2) select xl sheet
ws = wb['suqid game']

# 3) data Write
ws['A3'] = "2"
ws['B3'] = "강남사회서비스지원센터"
ws['C3'] = "02-3446-9982"

# 4) Saving data
wb.save(fpath)
