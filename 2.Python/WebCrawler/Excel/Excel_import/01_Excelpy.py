import openpyxl
import datetime

# 1) create xl
wb = openpyxl.Workbook()
now = datetime.date.today

# 2) create worksheet in xl
ws = wb.create_sheet('suqid game')

# 3) add data
ws['A1'] = 'No'
ws['B1'] = 'Name'
ws['C1'] = 'FaxNo.'

ws['A2'] = '1'
ws['B2'] = '강남복지재단'
ws['C2'] = '02-3446-1574'

# 4) save xl

wb.save(r"C:\Users\user\Documents\Study\2.Python\WebCrawler\Excel\data\Crawling_data.xlsx")
