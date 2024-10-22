import sys
from bs4 import BeautifulSoup as bs
from urllib.parse import quote_plus
from urllib.request import urlopen
import requests
import pandas as pd
import pyautogui
import datetime as dt
from datetime import datetime
import datetime
import time
from openpyxl import load_workbook
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from user_agent import generate_user_agent, generate_navigator
from cmd import PROMPT
from dataclasses import asdict
from gettext import find
from lib2to3.pgen2 import driver
from msilib.schema import File
from tkinter import BROWSE, BaseWidget, Button
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
<< << << < HEAD
== == == =
>>>>>> > 4b2de089e63da168dc4063e7e30b118736202878
<< << << < HEAD
== == == =
>>>>>> > 4b2de089e63da168dc4063e7e30b118736202878
# from PyQt5.QtWidgets import (
#     QApplication, QMainWindow, QLabel, QWidget, QLineEdit, QAction, qApp, QWidgetAction, QPushButton, QHBoxLayout, QVBoxLayout, QTextEdit)
# from PyQt5.QtCore import *
# from PyQt5.QtGui import *


# 0)class QtGUI(QWidget): 프로그램 위젯 UI/UX 구성.  //  UI 패키징 전체 보류.
# class MyApp(QMainWindow):
#
#    def __init__(self):
#        super().__init__()
#        self.initUI()
#        self.setWindowTitle("Web Crawler_업체정보수집기")
#
#    def initUI(self):
#        #
#
#        # 라벨 설정
#        self.lbl = QLabel(self)
#        self.lbl.move(60, 40)
#        qle = QLineEdit(self)
#        # 창 아이콘 설정
#        self.setWindowIcon(QIcon('Data\SANZIE_Fabicon.ico'))
#        qle.move(60, 100)
#        qle.textChanged[str].connect(self.onChanged)  # 데이터 입력받을 Textbox 생성
#        # 상태바 설정
#        SangTae = '대기'
#        self.statusBar().showMessage(SangTae)
#        self.setGeometry(300, 300, 300, 200)
#        self.show()
#
#    def onChanged(self, text):
#        self.lbl.setText(text)
#        self.lbl.adjustSize()
#
#
# if __name__ == '__main__':
#    app = QApplication(sys.argv)
#    ex = MyApp()
#    sys.exit(app.exec_())
#
#
# 0) pyautogui setup
# 프로그램 임시 사용을 위한 pyautogui 로 에러체크.
screenWidth, screenHeight = pyautogui.size()
screenWidth, screenHeight
(2560, 1440)


# 저장 영역 및 참조 데이터 불러오기(추후 수정)
fpath = "Crawling_data.xlsx"
dbpath = "sidogun.xlsx"

# 1) 저장영역 참조 데이터 변수선언
wb = openpyxl.load_workbook(fpath)
<< << << < HEAD
now = datetime.date.today
databook = openpyxl.load_workbook(dbpath)
databookWs = databook.active
crawling_data = wb.active

== == == =
databook = openpyxl.load_workbook(dbpath)
databook1 = databook["seoul"]
databook2 = databook["geyonggi"]
databook3 = databook["gangwon"]
databook4 = databook["jeonbuk"]
databook5 = databook["jeonnam"]
databook6 = databook["chungbuk"]
databook7 = databook["chungnam"]
databook8 = databook["geyongbuk"]
databook9 = databook["geyongnam"]
databook10 = databook["busan"]
databook11 = databook["jeju"]
databook20 = databook["allarea"]
crawling_data = wb.active
nowTime = datetime.now()
startTime = time.time()
>>>>>> > 4b2de089e63da168dc4063e7e30b118736202878
# 크롬 브라우저 패치, 설정변수 전역 초기화, 크롤링 세팅값 지정
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
<< << << < HEAD

service = Service(executable_path=ChromeDriverManager().install())
browser = webdriver.Chrome(service=service, options=chrome_options)


# 검색받을 데이터값 할당받기
keyword = pyautogui.prompt("검색어를 입력해 주세요")

== == == =
userAgent = generate_user_agent()
service = Service(executable_path=ChromeDriverManager().install())
browser = webdriver.Chrome(service=service, options=chrome_options)
chrome_options.add_argument(f'user-agent={userAgent}')
# 검색받을 데이터값 할당받기
keyword = pyautogui.prompt("검색어를 입력해 주세요")
findLocation = pyautogui.prompt(
    "검색받을 지역을 선택하세요.\n(1=서울,2=경기,3=강원,4=전북,5=전남,6=충북,7=충남,8=경북,9=경남,10=부산,11=제주,20=전 지역)\n * 주의 : 전 지역을 검색 할 경우 키워드에 따라 매우 많은 시간이 소요됩니다.")
print("크롤링을 시작합니다.")
print(nowTime.strftime('%Y-%m-%d %H:%M:%S'))
>>>>>> > 4b2de089e63da168dc4063e7e30b118736202878
# 화면 스크롤 초기값 할당
beforeH = browser.execute_script("return window.scrollY")

# 브라우저 오픈, 크롤링 시작점.
<< << << < HEAD
browser.get("https://map.naver.com/")
browser.implicitly_wait(1)

# 브라우저 서치 css값 search 변수에 할당
search = browser.find_element(
    By.CSS_SELECTOR, 'input.input_search')
== == == =
browser.implicitly_wait(100)
browser.get("https://www.useragentstring.com/")
userAgent = browser.find_element(By.ID, 'uas_textfeld').text
print('크롤링간 서버 차단 방지용 우회 UserAgent 추출중..')
print('우회용 User-Agent Key 추출 성공.')
print(userAgent)
browser.get("https://map.naver.com/")

# 브라우저 서치 css값 search 변수에 할당
search = browser.find_element(
    By.CSS_SELECTOR, '*.input_search')
>>>>>> > 4b2de089e63da168dc4063e7e30b118736202878
search.click()

# 브라우저 검색창 클리어 css값 clear 변수에 할당
# x 버튼 활성화
search.send_keys(" ")
clearSearch = browser.find_element(
    By.CSS_SELECTOR, 'button.button_clear')

# 데이터 크롤링 관련 지역 할당. 추후 frame 업데이트시에 버튼으로 지역별로 검색토록.
all_values = []
<< << << < HEAD
for row in databookWs.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
        all_values.append(row_value)
print(all_values)

# 크롤링 대상 접속 사이트 추가 팝업창 나타났을시의 제어용 변수 선언.
== == == =
if findLocation == "1":
    for row in databook1.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "2":
    for row in databook2.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "3":
    for row in databook3.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "4":
    for row in databook4.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "5":
    for row in databook5.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "6":
    for row in databook6.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "7":
    for row in databook7.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "8":
    for row in databook8.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "9":
    for row in databook9.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "10":
    for row in databook10.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "11":
    for row in databook11.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
elif findLocation == "20":
    for row in databook20.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
            all_values.append(row_value)
    print(all_values)
else:
    pyautogui.alert("검색받을 지역숫자를 잘못입력하였습니다.\n 다시 입력하시기 바랍니다.")

    # 크롤링 대상 접속 사이트 추가 팝업창 나타났을시의 제어용 변수 선언.
>>>>>> > 4b2de089e63da168dc4063e7e30b118736202878
tabs = browser.window_handles

# !--크롤링 시작점--!
i = 0
for crawling in all_values:
    browser.implicitly_wait(3)
    browser.switch_to.default_content()
    clearSearch = browser.find_element(
        By.CSS_SELECTOR, 'button.button_clear')
    clearSearch.click()
# 할당된 지역변수 값 순서대로 기입 후 해당 지역권으로 이동.
    search.send_keys(all_values[i])
    time.sleep(0.1)
    search.send_keys(Keys.ENTER)
    time.sleep(0.5)
    clearSearch = browser.find_element(
        By.CSS_SELECTOR, 'button.button_clear')
    clearSearch.click()
# 이동된 지역권에서 사용자에게 입력받은 키워드 값 검색.
    search.send_keys(keyword)
    time.sleep(0.1)
    search.send_keys(Keys.ENTER)
    time.sleep(0.5)
# 화면 좌측 업체 리스트로 프레임 포커스 전환.
    frame = browser.find_element(By.CSS_SELECTOR, "iframe#searchIframe")
    browser.switch_to.frame(frame)
    time.sleep(0.5)
    scroll_div = browser.find_element(
        By.XPATH, "/html/body/div[3]/div/div/div[1]")
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
<< << << < HEAD
    time.sleep(0.1)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.1)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.1)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.1)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.1)
== == == =
    time.sleep(0.2)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.2)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.2)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.2)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.2)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.2)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.2)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.2)
>>>>>> > 4b2de089e63da168dc4063e7e30b118736202878
    # window xposition 최하단으로 내려 div값 전체 나오도록 함
    final_result = []
    sort_result = []
    j = 2
    while j <= 10:
<< << << < HEAD
        stores = browser.find_elements(
            By.CSS_SELECTOR, "span.place_bluelink.YwYLL")
        print(stores)
        # 해당 페이지에서 표시된 모든 업체 정보를 stores 변수에 담은 후 각각의 업체정보 진입
        upChae = 1
        for store in stores:
            browser.implicitly_wait(1)
            try:
                name = store.find_element(
                    By.CSS_SELECTOR, "span.place_bluelink.YwYLL").text  # 업체명 크롤링 시작점
            except:
                name = store.find_element(
                    By.XPATH, f"/html/body/div[3]/div/div[2]/div[1]/ul/li[{upChae}]/div[1]/a/div/div/span[1]").text  # 업체명 크롤링 시작점
            click_name = store.find_element(
                By.CSS_SELECTOR, "div.div.C6RjW")
== == == =
        #########################################################################################################
        stores = browser.find_elements(
            By.CSS_SELECTOR, "span.YwYLL")
        print("검색된 페이지 내 자료의 갯수", len(stores))
# 네이버 크롤링 방지를 위해 CSS_SELECTOR 값 변화함(일정 주기마다) 찾아서 변경해야 함.
#########################################################################################################
        # 해당 페이지에서 표시된 모든 업체 정보를 stores 변수에 담은 후 각각의 업체정보 진입
        for store in stores:
            browser.implicitly_wait(5)
            time.sleep(1)
            try:
                name = store.text  # 업체명 크롤링 시작점
                click_name = store
            except:
                pass
>>>>>> > 4b2de089e63da168dc4063e7e30b118736202878
            try:
                # 너무 자세히 검색하여 결과값이 특정지어진 상태로 검색되었을 경우의 오류 처리.
                click_name.click()
            except:
                pass
            browser.switch_to.default_content()  # 브라우저 내부 세부 프레임 전환토록 default frame 전환
            # 내부 프레임 포커스를 못잡아 내서 안쪽 프레임 블럭을 못읽어냄. 따로 체크. 체크완료.
            frame_in = browser.find_element(By.ID, "entryIframe")
            browser.switch_to.frame(frame_in)  # 세부프레임으로 focus 진입
            time.sleep(0.5)
            try:
                com_address = browser.find_element(
                    By.XPATH, "//*[@id='app-root']/div/div/div/div/div/div/div/ul/li[1]/div/a/span[1]").text
            except:
                com_address = " "
            try:
<< << << < HEAD
== == == =
                com_phoneNum = browser.find_element(
                    By.CSS_SELECTOR, "span.dry01").text
            except:
                com_phoneNum = " "
            try:
>>>>>>> 4b2de089e63da168dc4063e7e30b118736202878
                click_url = browser.find_element(
                    By.XPATH, "//*[@id='app-root']/div/div/div/div/div/div/div/ul/li[6]/div/div/a")
                link_url = click_url.text
                # 사이트 접속하고 사이트 내 fax 값 찾아내는 시작점.
            except:
                try:
<<<<<<< HEAD
                    click_url = browser.find_element(
                        By.XPATH, "//*[@id='app-root']/div/div/div/div/div/div/div/ul/li/div/div/a")
                    link_url = click_url.text
=======
                    #########################################################################################################
                    click_url = browser.find_element(
                        By.CSS_SELECTOR, "a.JhzE0")
                    link_url = click_url.text
# 네이버 크롤링 방지를 위해 CSS_SELECTOR 값 변화함(일정 주기마다) 찾아서 변경해야 함.
#########################################################################################################
>>>>>>> 4b2de089e63da168dc4063e7e30b118736202878
                except:
                    link_url = " "
            if link_url != " ":
                click_url.click()
                while len(browser.window_handles) > 2:
<<<<<<< HEAD
=======
                    time.sleep(2)
                    browser.implicitly_wait(10)
>>>>>>> 4b2de089e63da168dc4063e7e30b118736202878
                    browser.switch_to.window(browser.window_handles[1])
                    browser.close()
                try:
                    time.sleep(0.5)
                    # 여러개의 팝업창이 나타났을 경우 메인 크롤링 페이지를 제외한 나머지 팝업페이지 닫기
                    # print(tabs)
                    browser.switch_to.window(browser.window_handles[1])
                    # 뷰티풀수프로 크롤링.
                    fax_no = browser.find_element(
                        By.XPATH, "//*[contains(text(),'팩스')]").text
                    if fax_no == '팩스':
                        fax_no_pe = browser.find_element(
                            By.XPATH, "//*[contains(text(),'팩스')]").find_element(By.XPATH, '..').text
                        fax_no = fax_no_pe
                    else:
                        pass
                except:
                    try:
                        fax_no = browser.find_element(
                            By.XPATH, "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZЙ', 'abcdefghijklmnopqrstuvwxyzй'),'fax')]").text
                    except:
                        try:
                            # 아예 안찾아졌을 경우 bs4 모듈 사용하여 html 내 존재하는 특정 문자열 대조하여 찾아냄.
                            page = requests.get(link_url)
                            soup = bs(page.text, "html.parser")
<<<<<<< HEAD
                            print(soup)
=======
                            # print(soup)
>>>>>>> 4b2de089e63da168dc4063e7e30b118736202878
                            fax_no = soup.find(text='fax')
                            if fax_no == 'None':
                                fax_no = soup.find(text='팩스')
                                if fax_no == 'None':
                                    fax_no = soup.find(text='FAX')
                        except:
                            fax_no = " "
                finally:
                    # 메인 프레임을 제외한 나머지 인터넷 탭 전체 닫아야 함 !!!!! 수정 필요.
                    # 수정함.
<<<<<<< HEAD
                    browser.close()
                    try:
                        browser.switch_to.window(browser.window_handles[0])
                    except:
                        try:
                            browser.switch_to.window(browser.window_handles[0])
                            time.sleep(3)
=======
                    if len(browser.window_handles) != 1:
                        # 웹상의 오류로 탭이 정상적으로 닫히지 않았는지 체크
                        browser.implicitly_wait(5)
                        browser.close()
                    try:
                        browser.switch_to.window(browser.window_handles[0])
                        browser.switch_to.default_content()
                        time.sleep(2)
                    except:
                        try:
                            time.sleep(3)
                            browser.switch_to.window(
                                browser.window_handles[0])
                            browser.switch_to.default_content()
>>>>>>> 4b2de089e63da168dc4063e7e30b118736202878
                        except:
                            pass
            else:
                fax_no = " "
<<<<<<< HEAD
            time.sleep(0.5)
            store_info = {
                '업체명': name,
                '주소': com_address,
                '홈페이지': link_url,
                '팩스번호': fax_no
            }
            # Crawling data 를 store_info 변수에 저장
            print(store_info)
            # 재정렬 된 자료를 final_result 변수에 담아냄. 추후 엑셀 또는 CSV 파일로 추출작업 필요함.

            time.sleep(2)
            sort_result.append(store_info)
            browser.switch_to.default_content()
            browser.switch_to.frame(frame)
            # 한 페이지 크롤링 끝
            upChae = upChae+1
=======
            try:
                time.sleep(0.5)
                store_info = {
                    '업체명': name,
                    '주소': com_address,
                    '전화번호': com_phoneNum,
                    '팩스번호': fax_no,
                    '홈페이지': link_url
                }
                print(store_info)
                # Crawling data 를 store_info 변수에 저장
                sort_result.append(store_info)
                browser.switch_to.window(browser.window_handles[0])
                browser.switch_to.default_content()
                frame_out = browser.find_element(
                    By.ID, "searchIframe")
                browser.switch_to.frame(frame_out)
            except:
                pass
            # 재정렬 된 자료를 final_result 변수에 담아냄. 추후 엑셀 또는 CSV 파일로 추출작업 필요함.
            # 한 페이지 크롤링 끝
            time.sleep(0.5)
>>>>>>> 4b2de089e63da168dc4063e7e30b118736202878
        # 다음 페이지 존재하는지 여부 파악 후 다음 페이지로 이동
        try:
            next_button = browser.find_element(By.LINK_TEXT, str(j))
            next_button.click()
            time.sleep(0.5)
        except:
            break
        j = j+1
<<<<<<< HEAD
    i = i+1
# 중복된 업체 정보를 제거하여 데이터 재정렬
final_result = pd.DataFrame(sort_result)

# 재정렬 된 자료 excel array 틀에 맞게 append후 data로 저장.
for r in dataframe_to_rows(final_result, index=True, header=True):
    crawling_data.append(r)
wb.save()
print(final_result)

# 데이터 저장, 불러오기 폴더 바꾸기. 아직안함.

# 배포용 프로그램으로 변환작업 필요함.
# 성 소수자와 젠더이슈 
=======
        time.sleep(2)
        browser.switch_to.window(browser.window_handles[0])
        browser.switch_to.default_content()
        frame_out = browser.find_element(
            By.ID, "searchIframe")
        browser.switch_to.frame(frame_out)
    # 중복된 업체 정보를 제거하여 데이터 재정렬
    final_result = pd.DataFrame(sort_result)
    # 재정렬 된 자료 excel array 틀에 맞게 append후 data로 저장.
    for r in dataframe_to_rows(final_result, index=True, header=True):
        crawling_data.append(r)
    i = i+1

wb.save('결과.xlsx')
print(final_result)

print("데이터 크롤링이 종료되었습니다.")
endtime = datetime.now()
print(endtime.strftime('%Y-%m-%d %H:%M:%S'))
sec = time.time()-startTime
times = str(dt.timedelta(seconds=sec))
short = times.split(".")[0]
print("추출된 데이터의 갯수 :", len(final_result))
print(f"추출간 소요시간      :{short} ")
print("")
pyautogui.alert(
    "크롤링이 완료되었습니다.\n 결과 파일에 정상적으로 저장이 되었는지 확인 하시기 바랍니다. \n 크롤링 도중 접근이 제한되어 비정상적으로 정지되었을 경우,\n 엑셀로 데이터가 이전되지 않을 수 있으니 \n 저장이 되지 않았을 경우 DOS 내의 데이터를 복사하십시오.")
# 데이터 저장, 불러오기 폴더 바꾸기. 아직안함.

# 배포용 프로그램으로 변환작업 필요함.
>>>>>>> 4b2de089e63da168dc4063e7e30b118736202878
