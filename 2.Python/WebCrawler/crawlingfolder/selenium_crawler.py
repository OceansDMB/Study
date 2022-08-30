from cmd import PROMPT
from dataclasses import asdict
from gettext import find
from lib2to3.pgen2 import driver
from msilib.schema import File
from tkinter import BROWSE, BaseWidget
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from openpyxl import load_workbook
import time
import datetime
import pyautogui
import pandas as pd
import requests
from urllib.request import urlopen
from urllib.parse import quote_plus
from bs4 import BeautifulSoup as bs

# 0) pyautogui setup
screenWidth, screenHeight = pyautogui.size()
screenWidth, screenHeight
(2560, 1440)


# 저장 영역 및 참조 데이터 불러오기(추후 수정)
fpath = r"C:\Users\user\Documents\Study\2.Python\WebCrawler\Excel\data\Crawling_data.xlsx"
dbpath = r"C:\Users\user\Documents\Study\2.Python\WebCrawler\Excel\data\sidogun.xlsx"

# 1) 저장영역 참조 데이터 변수선언
wb = openpyxl.load_workbook(fpath)
now = datetime.date.today
databook = openpyxl.load_workbook(dbpath)
databookWs = databook.active
crawling_data = wb.active

# 크롬 브라우저 패치, 설정변수 전역 초기화, 크롤링 세팅값 지정
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

service = Service(executable_path=ChromeDriverManager().install())
browser = webdriver.Chrome(service=service, options=chrome_options)


# 검색받을 데이터값 할당받기
keyword = pyautogui.prompt("검색어를 입력해 주세요")

# 화면 스크롤 초기값 할당
beforeH = browser.execute_script("return window.scrollY")

# 브라우저 오픈, 크롤링 시작점.
browser.get("https://map.naver.com/")
browser.implicitly_wait(1)

# 브라우저 서치 css값 search 변수에 할당
search = browser.find_element(
    By.CSS_SELECTOR, 'input.input_search')
search.click()

# 브라우저 검색창 클리어 css값 clear 변수에 할당
# x 버튼 활성화
search.send_keys(" ")
clearSearch = browser.find_element(
    By.CSS_SELECTOR, 'button.button_clear')

# 데이터 크롤링 관련 지역 할당. 추후 frame 업데이트시에 버튼으로 지역별로 검색토록.
all_values = []
for row in databookWs.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
        all_values.append(row_value)
print(all_values)

# 크롤링 대상 접속 사이트 추가 팝업창 나타났을시의 제어용 변수 선언.
tabs = browser.window_handles

# !--크롤링 시작점--!
i = 0
for crawling in all_values:
    browser.implicitly_wait(3)
    browser.switch_to.default_content()
    clearSearch = browser.find_element(
        By.CSS_SELECTOR, 'button.button_clear')
    clearSearch.click()
# 할당된 지역변수 값 순서대로 기입 후 해당 지역권으로 이동.
    search.send_keys(all_values[i])
    time.sleep(0.1)
    search.send_keys(Keys.ENTER)
    time.sleep(0.5)
    clearSearch = browser.find_element(
        By.CSS_SELECTOR, 'button.button_clear')
    clearSearch.click()
# 이동된 지역권에서 사용자에게 입력받은 키워드 값 검색.
    search.send_keys(keyword)
    time.sleep(0.1)
    search.send_keys(Keys.ENTER)
    time.sleep(0.5)
# 화면 좌측 업체 리스트로 프레임 포커스 전환.
    frame = browser.find_element(By.CSS_SELECTOR, "iframe#searchIframe")
    browser.switch_to.frame(frame)
    time.sleep(0.5)
    scroll_div = browser.find_element(
        By.XPATH, "/html/body/div[3]/div/div/div[1]")
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.1)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.1)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.1)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.1)
    browser.execute_script("arguments[0].scrollBy(0,2000)", scroll_div)
    time.sleep(0.1)
    # window xposition 최하단으로 내려 div값 전체 나오도록 함
    final_result = []
    sort_result = []
    j = 2
    while j <= 10:
        stores = browser.find_elements(
            By.XPATH, "/html/body/div[3]/div/div/div[1]/ul/li")
        # print(stores)
        # 해당 페이지에서 표시된 모든 업체 정보를 stores 변수에 담은 후 각각의 업체정보 진입
        upChae = 1
        for store in stores:
            browser.implicitly_wait(1)
            try:
                name = store.find_element(
                    By.CSS_SELECTOR, f"#_pcmap_list_scroll_container > ul > li:nth-child({upChae}) > * > a:nth-child(1) > div > div > span.place_bluelink").text  # 업체명 크롤링 시작점
            except:
                name = store.find_element(
                    By.XPATH, f"/html/body/div[3]/div/div[2]/div[1]/ul/li[{upChae}]/div[1]/a/div/div/span[1]").text  # 업체명 크롤링 시작점
            click_name = store.find_element(
                By.CSS_SELECTOR, f"#_pcmap_list_scroll_container > ul > li:nth-child({upChae}) > * > a:nth-child(1) > div > div > span.place_bluelink")
            try:
                # 너무 자세히 검색하여 결과값이 특정지어진 상태로 검색되었을 경우의 오류 처리.
                click_name.click()
            except:
                pass
            browser.switch_to.default_content()  # 브라우저 내부 세부 프레임 전환토록 default frame 전환
            # 내부 프레임 포커스를 못잡아 내서 안쪽 프레임 블럭을 못읽어냄. 따로 체크.
            frame_in = browser.find_element(By.ID, "entryIframe")
            browser.switch_to.frame(frame_in)  # 세부프레임으로 focus 진입
            time.sleep(0.5)
            try:
                com_address = browser.find_element(
                    By.XPATH, "//*[@id='app-root']/div/div/div/div/div/div/div/ul/li[1]/div/a/span[1]").text
            except:
                com_address = " "
            try:
                click_url = browser.find_element(
                    By.XPATH, "//*[@id='app-root']/div/div/div/div/div/div/div/ul/li[5]/div/div/a")
                link_url = click_url.text
                # 사이트 접속하고 사이트 내 fax 값 찾아내기
            except:
                try:
                    click_url = browser.find_element(
                        By.XPATH, "//*[@id='app-root']/div/div/div/div/div/div/div/ul/li/div/div/a")
                    link_url = click_url.text
                except:
                    link_url = " "
            if link_url != " ":
                click_url.click()
                while len(browser.window_handles) > 2:
                    browser.switch_to.window(browser.window_handles[1])
                    browser.close()
                try:
                    time.sleep(0.5)
                    # 여러개의 팝업창이 나타났을 경우 메인 크롤링 페이지를 제외한 나머지 팝업페이지 닫기
                    # print(tabs)
                    browser.switch_to.window(browser.window_handles[1])
                    # 뷰티풀수프로 크롤링.
                    fax_no = browser.find_element(
                        By.XPATH, "//*[contains(text(),'팩스')]").text
                    if fax_no == '팩스':
                        fax_no_pe = browser.find_element(
                            By.XPATH, "//*[contains(text(),'팩스')]").find_element(By.XPATH, '..').text
                        fax_no = fax_no_pe
                    else:
                        pass
                except:
                    try:
                        fax_no = browser.find_element(
                            By.XPATH, "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZЙ', 'abcdefghijklmnopqrstuvwxyzй'),'fax')]").text
                    except:
                        try:
                            # 아예 안찾아졌을 경우 bs4 모듈 사용하여 html 내 존재하는 특정 문자열 대조하여 찾아냄.
                            page = requests.get(link_url)
                            soup = bs(page.text, "html.parser")
                            print(soup)
                            fax_no = soup.find(text='fax')
                            if fax_no == 'None':
                                fax_no = soup.find(text='팩스')
                                if fax_no == 'None':
                                    fax_no = soup.find(text='FAX')
                        except:
                            fax_no = " "
                finally:
                    # 메인 프레임을 제외한 나머지 인터넷 탭 전체 닫아야 함 !!!!! 수정 필요.
                    browser.close()
                    try:
                        browser.switch_to.window(browser.window_handles[0])
                    except:
                        try:
                            browser.switch_to.window(browser.window_handles[0])
                            time.sleep(3)
                        except:
                            break
            else:
                fax_no = " "
            time.sleep(0.5)
            store_info = {
                '업체명': name,
                '주소': com_address,
                '홈페이지': link_url,
                '팩스번호': fax_no
            }
            # Crawling data 를 store_info 변수에 저장
            print(store_info)
            # 재정렬 된 자료를 final_result 변수에 담아냄. 추후 엑셀 또는 CSV 파일로 추출작업 필요함.
            sort_result.append(store_info)
            browser.switch_to.default_content()
            browser.switch_to.frame(frame)
            time.sleep(0.5)
            # 한 페이지 크롤링 끝
            upChae = upChae+1
        # 다음 페이지 존재하는지 여부 파악 후 다음 페이지로 이동
        try:
            next_button = browser.find_element(By.LINK_TEXT, str(j))
            next_button.click()
            time.sleep(0.5)
        except:
            break
        j = j+1
    i = i+1
# 중복된 업체 정보를 제거하여 데이터 재정렬

final_result = pd.DataFrame(sort_result)

for r in dataframe_to_rows(final_result, index=True, header=True):
    crawling_data.append(r)
wb.save(r"C:\Users\user\Documents\Study\2.Python\WebCrawler\Excel\data\Crawling_data2.xlsx")
print(final_result)


# 데이터 저장, 불러오기 폴더 바꾸기.
