from lib2to3.pgen2 import driver
from tkinter import BROWSE
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl import load_workbook
import time
import datetime
import pyautogui

# 저장 영역 및 참조 데이터 불러오기(추후 수정)
fpath = r"C:\Users\user\Documents\Study\2.Python\WebCrawler\Excel\data\Crawling_data.xlsx"
dbpath = r"C:\Users\user\Documents\Study\2.Python\WebCrawler\Excel\data\sidogun.xlsx"

# 1) 저장영역 참조 데이터 변수선언
wb = openpyxl.load_workbook(fpath)
now = datetime.date.today
databook = openpyxl.load_workbook(dbpath)
databookWs = databook.active


# 크롬 브라우저 패치, 설정변수 전역 초기화, 크롤링 세팅값 지정
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

service = Service(executable_path=ChromeDriverManager().install())
browser = webdriver.Chrome(service=service, options=chrome_options)


# 검색받을 데이터값 할당받기
keyword = pyautogui.prompt("검색어를 입력해 주세요")

# 화면 스크롤 초기값 할당
beforeH = browser.execute_script("return window.scrollY")

# 브라우저 오픈, 크롤링 시작점.
browser.get("https://map.naver.com/")
browser.implicitly_wait(1)

search = browser.find_element(
    By.CSS_SELECTOR, 'button.button_search')
search.click()

# 데이터 크롤링 관련 지역 할당.
all_values = []
for row in databookWs.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
        all_values.append(row_value)
print(all_values)
i = 0
time.sleep(1)
browser.implicitly_wait(1)
search.send_keys(all_values[1])
time.sleep(1)
search.send_keys(Keys.ENTER)
time.sleep(1)
search.send_keys(Keys.BACK_SPACE)
time.sleep(1)
search.send_keys(" "+keyword)
time.sleep(1)
search.send_keys(Keys.ENTER)
print(keyword)

while (True):
    pass
